from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.models.value_objects import ColumnDefinition, StylingConfig
from app.domain.repositories.generator import IFileGenerate


class ExcelGenerator(IFileGenerate):
    async def generate(
        self,
        data: list[dict[str, Any]],
        columns: list[ColumnDefinition],
        styling: StylingConfig | None = None,
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        # 1. Заголовки
        self._write_headers(ws, columns, styling)

        # 2. Данные
        self._write_data(ws, data, columns)

        # 3. Автоподбор ширины колонок
        self._auto_fit_columns(ws, columns)

        # 4. Сохранение в bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _to_argb_color(hex_color: str) -> str:
        """Преобразует RRGGBB в FFRRGGBB."""
        color = hex_color.lstrip("#")
        if len(color) == 6:
            result = "FF" + color
            return result
        return color

    def _write_headers(
        self,
        ws: Worksheet,
        columns: list[ColumnDefinition],
        styling: StylingConfig | None = None,
    ) -> None:
        """Запись заголовков колонок"""

        for col_indx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_indx, value=col.header)

            if styling:
                if styling.header_bg_color:
                    bg_color = self._to_argb_color(styling.header_bg_color)
                    cell.fill = PatternFill(
                        start_color=bg_color,
                        end_color=bg_color,
                        fill_type="solid",
                    )

                if styling.header_font_color:
                    font_color = self._to_argb_color(styling.header_font_color)
                    cell.font = Font(color=font_color, bold=True)
                else:
                    cell.font = Font(bold=True)

                if styling.font_size:
                    cell.font = Font(size=styling.font_size, bold=True)

                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    def _write_data(
        self, ws: Worksheet, data: list[dict[str, Any]], columns: list[ColumnDefinition]
    ) -> None:
        """Записывает строки с данными"""
        for row_indx, row_data in enumerate(data, start=2):
            for col_indx, col in enumerate(columns, start=1):
                value = row_data.get(col.field, "")
                cell = ws.cell(row=row_indx, column=col_indx, value=value)

                if col.format and isinstance(value, (int, float)):
                    cell.number_format = col.format

    def _auto_fit_columns(self, ws: Worksheet, columns: list[ColumnDefinition]) -> None:
        """Автоподбор ширины колонок"""
        for col_indx, col in enumerate(columns, start=1):
            if col.width:
                ws.column_dimensions[get_column_letter(col_indx)].width = col.width

            else:
                max_lenght = len(col.header)
                for row in ws.iter_rows(
                    min_row=2, max_col=col_indx, max_row=ws.max_row
                ):
                    for cell in row:
                        if cell.value:
                            max_lenght = max(max_lenght, len(str(cell.value)))

                adjusted_width = min(max_lenght + 2, 50)
                ws.column_dimensions[get_column_letter(col_indx)].width = adjusted_width
